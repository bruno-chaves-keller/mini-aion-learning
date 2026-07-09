from openpyxl import load_workbook

def read_excel(file_path):

    workbook = load_workbook(file_path)
    sheet = workbook.active

    model_data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        field = row[0]
        value = row[1]
        notes = row[2]

        model_data[field] = {
            "value": value,
            "notes": notes
        }

    return model_data


